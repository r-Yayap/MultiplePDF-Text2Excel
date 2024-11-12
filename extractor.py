#extractor.py

import shutil
import os
import multiprocessing
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import pymupdf as fitz
from utils import adjust_coordinates_for_rotation, find_tessdata
import re
from openpyxl.styles import Font



class TextExtractor:
    def __init__(self, pdf_folder, output_excel_path, areas, ocr_settings, include_subfolders):
        self.pdf_folder = pdf_folder
        self.output_excel_path = output_excel_path
        self.areas = areas
        self.ocr_settings = ocr_settings
        self.include_subfolders = include_subfolders
        self.tessdata_folder = find_tessdata() if ocr_settings["enable_ocr"] != "Off" else None
        self.temp_image_folder = "temp_images"
        self.headers = ["Size (Bytes)", "Date Last Modified", "Folder", "Filename", "Page No"] + \
                       [f"{area['title']}" if "title" in area else f"Area {i + 1}" for i, area in enumerate(self.areas)]

        if not os.path.exists(self.temp_image_folder):
            os.makedirs(self.temp_image_folder)

    def clean_text(self, text):
        """Cleans text by replacing newlines, stripping, and removing illegal characters."""
        replacement_char = '■'  # Character to replace prohibited control characters

        # Step 1: Replace newline and carriage return characters with a space
        text = text.replace('\n', ' ').replace('\r', ' ')

        # Step 2: Strip leading and trailing whitespace
        text = text.strip()

        # Step 3: Replace prohibited control characters with a replacement character
        text = re.sub(r'[\x00-\x1F\x7F-\x9F]', replacement_char, text)

        # Step 4: Remove extra spaces between words
        return re.sub(r'\s+', ' ', text)

    def start_extraction(self, progress_list, total_files):
        """Starts the extraction process using multiprocessing with progress tracking."""
        try:
            with multiprocessing.Pool() as pool:
                # Gather all PDF files in the specified folder
                pdf_files = self.get_pdf_files()
                total_files.value = len(pdf_files)  # Set total files count

                # Process each PDF file using multiprocessing with progress tracking
                results = pool.starmap(self.process_single_pdf,
                                       [(pdf_path, progress_list, total_files) for pdf_path in pdf_files])

                # Consolidate results into an Excel file after extraction is complete
                self.consolidate_results(results)
        except Exception as e:
            print(f"Error during extraction: {e}")

    def process_single_pdf(self, pdf_path, progress_list, total_files):
        """Processes a single PDF file, extracting text and images as necessary."""
        try:
            pdf_document = fitz.open(pdf_path)
            file_size = os.path.getsize(pdf_path)
            last_modified_date = datetime.fromtimestamp(os.path.getmtime(pdf_path)).strftime('%Y-%m-%d %H:%M:%S')
            folder = os.path.relpath(os.path.dirname(pdf_path), self.pdf_folder)
            filename = os.path.basename(pdf_path)

            result_rows = []  # Collect each page's data here

            for page_number in range(pdf_document.page_count):
                page = pdf_document[page_number]
                extracted_areas = []

                for area_index, area in enumerate(self.areas):
                    coordinates = area["coordinates"]
                    text_area, img_path = self.extract_text_from_area(page, coordinates, pdf_path, page_number,
                                                                      area_index)
                    extracted_areas.append((text_area or "", img_path))

                # Append all relevant data for this page as a row in results
                result_rows.append([file_size, last_modified_date, folder, filename, page_number + 1, extracted_areas])

            pdf_document.close()

            # Add each page to results
            progress_list.append(result_rows)
            return result_rows

        except Exception as e:
            print(f"Error processing {pdf_path}: {e}")
            return None




    def get_pdf_files(self):
        """Gathers all PDF files within the specified folder."""
        pdf_files = []
        for root_folder, subfolders, files in os.walk(self.pdf_folder):
            if not self.include_subfolders:
                subfolders.clear()
            pdf_files.extend(
                [os.path.join(root_folder, f) for f in files if f.lower().endswith('.pdf')]
            )
        return pdf_files



    def extract_text_from_area(self, page, area_coordinates, pdf_path, page_number, area_index):
        """Extracts text or image content from a specified area in a PDF page."""
        adjusted_coordinates = adjust_coordinates_for_rotation(
            area_coordinates, page.rotation, page.rect.height, page.rect.width
        )
        text_area = ""
        img_path = None

        try:
            # Text extraction without OCR
            if self.ocr_settings["enable_ocr"] == "Off":
                text_area = page.get_text("text", clip=adjusted_coordinates)

            # OCR only if there's no text
            elif self.ocr_settings["enable_ocr"] == "Text-first":
                text_area = page.get_text("text", clip=adjusted_coordinates)
                if not text_area.strip():
                    text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)

            # OCR for all areas, no images saved
            elif self.ocr_settings["enable_ocr"] == "OCR-All":
                text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)

            # OCR with image saving regardless of extracted text for Text1st+Image-beta
            elif self.ocr_settings["enable_ocr"] == "Text1st+Image-beta":
                text_area = page.get_text("text", clip=adjusted_coordinates)
                pix = page.get_pixmap(clip=area_coordinates, dpi=self.ocr_settings.get("dpi_value", 150))
                img_path = os.path.join(self.temp_image_folder,
                                        f"{os.path.basename(pdf_path)}_page{page_number + 1}_area{area_index}.png")
                pix.save(img_path)

                # Apply OCR if no text found
                if not text_area.strip():
                    text_area, _ = self.apply_ocr(page, area_coordinates, pdf_path, page_number, area_index)

            # Clean the extracted text
            text_area = self.clean_text(text_area)


        except fitz.EmptyFileError as e:
            print(f"Error: {pdf_path} is empty or corrupted: {e}")
            text_area = "Corrupted File"

        except fitz.FileNotFoundError as e:
            print(f"Error: {pdf_path} not found: {e}")
            text_area = "File Not Found"

        except RuntimeError as e:
            print(f"Runtime error on page {page_number + 1} in {pdf_path}: {e}")
            text_area = "Error Loading Page"

        except Exception as e:
            print(f"Unexpected error extracting text from area {area_index} in {pdf_path}, Page {page_number + 1}: {e}")
            text_area = "Extraction Error"

        return text_area, img_path

    def apply_ocr(self, page, coordinates, pdf_path, page_number, area_index):
        """Applies OCR on a specified area and returns the extracted text and image path."""
        if not self.tessdata_folder:
            print("Tessdata folder not found. OCR cannot proceed.")
            return "", ""

        pix = page.get_pixmap(clip=coordinates, dpi=self.ocr_settings.get("dpi_value", 150))
        pdfdata = pix.pdfocr_tobytes(language="eng", tessdata=self.tessdata_folder)
        clipdoc = fitz.open("pdf", pdfdata)
        text_area = "_OCR_" + clipdoc[0].get_text()
        img_path = os.path.join(self.temp_image_folder,
                                f"{os.path.basename(pdf_path)}_page{page_number + 1}_area{area_index}.png")
        pix.save(img_path)
        return text_area, img_path

    def consolidate_results(self, results):
        """Consolidates all extracted results into an Excel file with each page as a row."""
        wb = Workbook()
        ws = wb.active
        ws.append(self.headers)  # Add headers to the first row

        for result_pages in results:
            for row_data in result_pages:
                try:
                    # Unpack metadata and extracted area data for the page
                    file_size, last_modified, folder, filename, page_number, extracted_areas = row_data

                    # Prepare row for insertion
                    row_index = ws.max_row + 1
                    metadata = [file_size, last_modified, folder, filename, page_number]

                    # Write metadata to columns 1-5
                    for col, data in enumerate(metadata, start=1):
                        cell = ws.cell(row=row_index, column=col)
                        cell.value = data
                        if col == 4:  # Set filename as a hyperlink in column 4
                            cell.hyperlink = Hyperlink(
                                target=f"file://{os.path.abspath(os.path.join(self.pdf_folder, folder, filename))}",
                                ref=cell.coordinate)
                            cell.style = "Hyperlink"

                    # Write extracted areas to columns starting from 6
                    for i, (text, img_path) in enumerate(extracted_areas):
                        # Find the appropriate column index for the title assigned to this area
                        column_title = self.areas[i].get("title", f"Area {i + 1}")
                        col_index = self.headers.index(column_title) + 1  # Excel columns are 1-based

                        # Place text in the designated column, checking for OCR and cleaning
                        text_cell = ws.cell(row=row_index, column=col_index)
                        cleaned_text = self.clean_text(text.replace("_OCR_", ""))
                        text_cell.value = cleaned_text

                        # Highlight OCR text in red if "_OCR_" was detected
                        if "_OCR_" in text:
                            text_cell.font = Font(color="FF3300")

                        # If there is an associated image, add it to the cell
                        if img_path:
                            img = ExcelImage(img_path)
                            img.anchor = f"{get_column_letter(col_index)}{row_index}"
                            ws.add_image(img)

                except Exception as e:
                    print(f"Unexpected error consolidating data for {filename}: {e}")
                    ws.append([folder, filename, "Error"] + [""] * len(self.areas))

        # Save to the output path
        try:
            wb.save(self.output_excel_path)
            print(f"Consolidated results saved to {self.output_excel_path}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

        # Cleanup temporary images
        if os.path.exists(self.temp_image_folder):
            shutil.rmtree(self.temp_image_folder)

